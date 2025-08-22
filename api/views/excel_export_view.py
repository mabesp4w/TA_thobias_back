# views/excel_export_view.py
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import Cell
from rest_framework.decorators import action
from datetime import datetime
from django.db.models import Sum, Count, Avg, Q
from rest_framework.response import Response
from rest_framework import status
from api.serializers import ProdukTerjualSerializer
from crud.models import ProdukTerjual
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from crud.models import ProfilUMKM  # Hapus import User, hanya gunakan ProfilUMKM


class ExcelExportMixin:
    def format_rupiah(self, value):
        """Format number to Rupiah currency format"""
        if value is None or value == 0:
            return "Rp 0"
        # Convert to string and reverse
        str_value = str(int(value))[::-1]
        # Add dots every 3 digits
        groups = [str_value[i:i + 3] for i in range(0, len(str_value), 3)]
        # Join and reverse back
        formatted = '.'.join(groups)[::-1]
        return f"Rp {formatted}"

    def apply_rupiah_format(self, ws, cell):
        """Apply Rupiah number format to a cell"""
        cell.number_format = '#,##0'
        cell.value = f'Rp {cell.value:,.0f}' if isinstance(cell.value, (int, float)) else cell.value

    @action(detail=False, methods=['get'])
    def export_sales_report(self, request):
        """Export laporan penjualan ke Excel"""
        # Check if user is admin or UMKM
        if request.user.is_staff:
            # Admin can see all sales
            base_queryset = ProdukTerjual.objects.all()
        elif hasattr(request.user, 'role') and request.user.role == 'umkm':
            # UMKM can only see their own sales
            base_queryset = ProdukTerjual.objects.filter(produk__umkm=request.user)
        else:
            return Response({
                'status': 'error',
                'message': 'Anda tidak memiliki akses untuk mengexport laporan penjualan'
            }, status=status.HTTP_403_FORBIDDEN)

        # Tambahkan filter UMKM jika disediakan (hanya untuk admin)
        umkm_id = request.GET.get('umkm_id')
        if umkm_id and request.user.is_staff:
            try:
                profil_umkm = ProfilUMKM.objects.get(id=umkm_id)
                base_queryset = base_queryset.filter(produk__umkm=profil_umkm.user)
            except ProfilUMKM.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'UMKM tidak ditemukan'
                }, status=status.HTTP_404_NOT_FOUND)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Penjualan"
        # Headers
        headers = [
            'No', 'Tanggal Penjualan', 'Nama Produk', 'Kategori',
            'Jumlah Terjual', 'Satuan', 'Harga Jual', 'Total Penjualan',
            'Lokasi Penjualan', 'Kecamatan', 'Kabupaten', 'Catatan'
        ]
        # Add UMKM column if admin and no specific UMKM selected
        if request.user.is_staff and not umkm_id:
            headers.insert(3, 'UMKM')
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Get data with filters
        queryset = base_queryset.select_related(
            'produk',
            'produk__kategori',
            'produk__umkm',
            'produk__umkm__profil_umkm',
            'lokasi_penjualan',
            'lokasi_penjualan__kecamatan',
            'lokasi_penjualan__kecamatan__kabupaten'
        )
        # Apply date filters if provided
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                queryset = queryset.filter(tgl_penjualan__gte=start_date)
            except ValueError:
                pass
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                queryset = queryset.filter(tgl_penjualan__lte=end_date)
            except ValueError:
                pass
        # Apply additional filters from the viewset if available
        if hasattr(self, 'filter_queryset'):
            queryset = self.filter_queryset(queryset)
        # Order by date descending
        queryset = queryset.order_by('-tgl_penjualan')
        # Write data
        row_num = 2
        for idx, sale in enumerate(queryset, 1):
            col_idx = 1
            ws.cell(row=row_num, column=col_idx, value=idx)
            col_idx += 1
            ws.cell(row=row_num, column=col_idx,
                    value=sale.tgl_penjualan.strftime('%d/%m/%Y') if sale.tgl_penjualan else '-')
            col_idx += 1
            # Add UMKM name if admin and no specific UMKM
            if request.user.is_staff and not umkm_id:
                umkm_name = sale.produk.umkm.profil_umkm.nm_bisnis if sale.produk and sale.produk.umkm and sale.produk.umkm.profil_umkm else '-'
                ws.cell(row=row_num, column=col_idx, value=umkm_name)
                col_idx += 1
            ws.cell(row=row_num, column=col_idx, value=sale.produk.nm_produk if sale.produk else '-')
            col_idx += 1
            ws.cell(row=row_num, column=col_idx,
                    value=sale.produk.kategori.nm_kategori if sale.produk and sale.produk.kategori else '-')
            col_idx += 1
            ws.cell(row=row_num, column=col_idx, value=sale.jumlah_terjual or 0)
            col_idx += 1
            ws.cell(row=row_num, column=col_idx, value=sale.produk.satuan if sale.produk else '-')
            col_idx += 1
            # Harga Jual dengan format Rupiah
            harga_cell = ws.cell(row=row_num, column=col_idx, value=sale.harga_jual or 0)
            harga_cell.number_format = '"Rp "#,##0'
            col_idx += 1
            # Total Penjualan dengan format Rupiah
            total_cell = ws.cell(row=row_num, column=col_idx, value=sale.total_penjualan or 0)
            total_cell.number_format = '"Rp "#,##0'
            col_idx += 1
            ws.cell(row=row_num, column=col_idx,
                    value=sale.lokasi_penjualan.nm_lokasi if sale.lokasi_penjualan else "-")
            col_idx += 1
            ws.cell(row=row_num, column=col_idx,
                    value=sale.lokasi_penjualan.kecamatan.nm_kecamatan if sale.lokasi_penjualan and sale.lokasi_penjualan.kecamatan else "-")
            col_idx += 1
            ws.cell(row=row_num, column=col_idx,
                    value=sale.lokasi_penjualan.kecamatan.kabupaten.nm_kabupaten if sale.lokasi_penjualan and sale.lokasi_penjualan.kecamatan and sale.lokasi_penjualan.kecamatan.kabupaten else "-")
            col_idx += 1
            ws.cell(row=row_num, column=col_idx, value=sale.catatan or "-")
            row_num += 1
        # Auto-adjust columns width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        # Add summary row if there's data
        if queryset.exists():
            summary_row = row_num + 1
            ws.cell(row=summary_row, column=1, value="TOTAL")
            # Find the correct column for total based on whether UMKM column exists
            total_col = 8 if (request.user.is_staff and not umkm_id) else 7
            total_cell = ws.cell(row=summary_row, column=total_col + 1,
                                 value=f"=SUM({chr(64 + total_col + 1)}2:{chr(64 + total_col + 1)}{row_num - 1})")
            total_cell.number_format = '"Rp "#,##0'
            # Style summary row
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=summary_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        # Save to response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Create filename with date range info
        filename_parts = ['laporan_penjualan']
        if request.user.is_staff:
            if umkm_id:
                filename_parts.append('umkm_spesifik')
            else:
                filename_parts.append('all_umkm')
        if start_date:
            filename_parts.append(f'dari_{start_date.strftime("%Y%m%d")}')
        if end_date:
            filename_parts.append(f'sampai_{end_date.strftime("%Y%m%d")}')
        filename_parts.append(datetime.now().strftime("%Y%m%d_%H%M%S"))
        filename = '_'.join(filename_parts) + '.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'])
    def export_sales_analysis(self, request):
        """Export analisis penjualan dengan multiple sheets"""
        wb = Workbook()
        # Get base queryset based on user role
        if request.user.is_staff:
            base_queryset = ProdukTerjual.objects.all()
            analysis_title = 'LAPORAN ANALISIS PENJUALAN - SEMUA UMKM'
        elif hasattr(request.user, 'role') and request.user.role == 'umkm':
            base_queryset = ProdukTerjual.objects.filter(produk__umkm=request.user)
            analysis_title = f'LAPORAN ANALISIS PENJUALAN - {request.user.get_full_name()}'
        else:
            return Response({
                'status': 'error',
                'message': 'Anda tidak memiliki akses untuk mengexport analisis penjualan'
            }, status=status.HTTP_403_FORBIDDEN)

        # Tambahkan filter UMKM jika disediakan (hanya untuk admin)
        umkm_id = request.GET.get('umkm_id')
        if umkm_id and request.user.is_staff:
            try:
                profil_umkm = ProfilUMKM.objects.get(id=umkm_id)
                base_queryset = base_queryset.filter(produk__umkm=profil_umkm.user)
                analysis_title = f'LAPORAN ANALISIS PENJUALAN - {profil_umkm.nm_bisnis}'
            except ProfilUMKM.DoesNotExist:
                return Response({
                    'status': 'error',
                    'message': 'UMKM tidak ditemukan'
                }, status=status.HTTP_404_NOT_FOUND)

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Ringkasan"
        # Summary metrics
        total_sales = base_queryset.aggregate(
            total_revenue=Sum('total_penjualan'),
            total_quantity=Sum('jumlah_terjual'),
            total_transactions=Count('id'),
            avg_transaction=Avg('total_penjualan')
        )
        # Style for headers
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # Write summary
        ws1['A1'] = analysis_title
        ws1['A1'].font = Font(bold=True, size=16)
        ws1.merge_cells('A1:D1')
        ws1['A3'] = 'Total Pendapatan:'
        revenue_cell = ws1['B3']
        revenue_cell.value = total_sales['total_revenue'] or 0
        revenue_cell.number_format = '"Rp "#,##0'
        ws1['A4'] = 'Total Produk Terjual:'
        ws1['B4'] = total_sales['total_quantity'] or 0
        ws1['A5'] = 'Total Transaksi:'
        ws1['B5'] = total_sales['total_transactions'] or 0
        ws1['A6'] = 'Rata-rata per Transaksi:'
        avg_cell = ws1['B6']
        avg_cell.value = total_sales['avg_transaction'] or 0
        avg_cell.number_format = '"Rp "#,##0'
        # Add generated date
        ws1['A8'] = 'Tanggal Generate:'
        ws1['B8'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        # Auto-adjust columns
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        # Sheet 2: Sales by Product
        ws2 = wb.create_sheet("Penjualan per Produk")
        # Product sales data - include UMKM info if admin and no specific UMKM
        if request.user.is_staff and not umkm_id:
            product_sales = base_queryset.values(
                'produk__nm_produk',
                'produk__kategori__nm_kategori',
                'produk__umkm__profil_umkm__nm_bisnis'
            ).annotate(
                total_quantity=Sum('jumlah_terjual'),
                total_revenue=Sum('total_penjualan'),
                transaction_count=Count('id')
            ).order_by('-total_revenue')
            # Headers for product sheet
            product_headers = ['Nama Produk', 'Kategori', 'UMKM', 'Jumlah Terjual', 'Total Pendapatan',
                               'Jumlah Transaksi']
        else:
            product_sales = base_queryset.values(
                'produk__nm_produk',
                'produk__kategori__nm_kategori'
            ).annotate(
                total_quantity=Sum('jumlah_terjual'),
                total_revenue=Sum('total_penjualan'),
                transaction_count=Count('id')
            ).order_by('-total_revenue')
            # Headers for product sheet
            product_headers = ['Nama Produk', 'Kategori', 'Jumlah Terjual', 'Total Pendapatan', 'Jumlah Transaksi']
        for col, header in enumerate(product_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # Product data
        for idx, product in enumerate(product_sales, 2):
            col_idx = 1
            ws2.cell(row=idx, column=col_idx, value=product['produk__nm_produk'] or '-')
            col_idx += 1
            ws2.cell(row=idx, column=col_idx, value=product['produk__kategori__nm_kategori'] or '-')
            col_idx += 1
            if request.user.is_staff and not umkm_id:
                umkm_name = product.get('produk__umkm__profil_umkm__nm_bisnis', '-')
                ws2.cell(row=idx, column=col_idx, value=umkm_name)
                col_idx += 1
            ws2.cell(row=idx, column=col_idx, value=product['total_quantity'] or 0)
            col_idx += 1
            revenue_cell = ws2.cell(row=idx, column=col_idx, value=product['total_revenue'] or 0)
            revenue_cell.number_format = '"Rp "#,##0'
            col_idx += 1
            ws2.cell(row=idx, column=col_idx, value=product['transaction_count'] or 0)
        # Auto-adjust columns for product sheet
        for col in range(1, len(product_headers) + 1):
            ws2.column_dimensions[chr(64 + col)].width = 20
        # Sheet 3: Sales by Location
        ws3 = wb.create_sheet("Penjualan per Lokasi")
        # Location sales data
        location_sales = base_queryset.exclude(
            lokasi_penjualan__isnull=True
        ).values(
            'lokasi_penjualan__nm_lokasi',
            'lokasi_penjualan__kecamatan__nm_kecamatan',
            'lokasi_penjualan__kecamatan__kabupaten__nm_kabupaten'
        ).annotate(
            total_revenue=Sum('total_penjualan'),
            transaction_count=Count('id'),
            total_quantity=Sum('jumlah_terjual')
        ).order_by('-total_revenue')
        # Headers for location sheet
        location_headers = ['Lokasi', 'Kecamatan', 'Kabupaten', 'Total Pendapatan', 'Jumlah Transaksi',
                            'Total Quantity']
        for col, header in enumerate(location_headers, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # Location data
        for idx, location in enumerate(location_sales, 2):
            ws3.cell(row=idx, column=1, value=location['lokasi_penjualan__nm_lokasi'] or '-')
            ws3.cell(row=idx, column=2, value=location['lokasi_penjualan__kecamatan__nm_kecamatan'] or '-')
            ws3.cell(row=idx, column=3, value=location['lokasi_penjualan__kecamatan__kabupaten__nm_kabupaten'] or '-')
            revenue_cell = ws3.cell(row=idx, column=4, value=location['total_revenue'] or 0)
            revenue_cell.number_format = '"Rp "#,##0'
            ws3.cell(row=idx, column=5, value=location['transaction_count'] or 0)
            ws3.cell(row=idx, column=6, value=location['total_quantity'] or 0)
        # Auto-adjust columns for location sheet
        for col in range(1, 7):
            ws3.column_dimensions[chr(64 + col)].width = 18
        # Sheet 4: Monthly Sales Trend
        ws4 = wb.create_sheet("Trend Bulanan")
        # Fix for PostgreSQL date formatting
        from django.db import connection
        if 'postgresql' in connection.vendor:
            # PostgreSQL query
            monthly_sales = base_queryset.extra(
                select={'month': "TO_CHAR(tgl_penjualan, 'YYYY-MM')"}
            ).values('month').annotate(
                total_revenue=Sum('total_penjualan'),
                transaction_count=Count('id')
            ).order_by('month')
        else:
            # MySQL query
            monthly_sales = base_queryset.extra(
                select={'month': "DATE_FORMAT(tgl_penjualan, '%%Y-%%m')"}
            ).values('month').annotate(
                total_revenue=Sum('total_penjualan'),
                transaction_count=Count('id')
            ).order_by('month')
        # Headers for monthly trend
        monthly_headers = ['Bulan', 'Total Pendapatan', 'Jumlah Transaksi']
        for col, header in enumerate(monthly_headers, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        # Monthly data
        for idx, monthly in enumerate(monthly_sales, 2):
            # Format month for better readability
            try:
                month_obj = datetime.strptime(monthly['month'], '%Y-%m')
                formatted_month = month_obj.strftime('%B %Y')
            except:
                formatted_month = monthly['month']
            ws4.cell(row=idx, column=1, value=formatted_month)
            revenue_cell = ws4.cell(row=idx, column=2, value=monthly['total_revenue'] or 0)
            revenue_cell.number_format = '"Rp "#,##0'
            ws4.cell(row=idx, column=3, value=monthly['transaction_count'] or 0)
        # Auto-adjust columns for monthly sheet
        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 18
        ws4.column_dimensions['C'].width = 18
        # Save
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        if request.user.is_staff:
            if umkm_id:
                filename = f'analisis_penjualan_umkm_{umkm_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            else:
                filename = f'analisis_penjualan_all_umkm_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        else:
            filename = f'analisis_penjualan_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response


class SalesViewSet(ExcelExportMixin, viewsets.GenericViewSet):
    serializer_class = ProdukTerjualSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Get queryset based on user role
        Admin can see all sales, UMKM can only see their own
        """
        if self.request.user.is_staff:
            # Admin can see all sales
            return ProdukTerjual.objects.all().select_related(
                'produk',
                'produk__kategori',
                'produk__umkm',
                'produk__umkm__profil_umkm',
                'lokasi_penjualan',
                'lokasi_penjualan__kecamatan',
                'lokasi_penjualan__kecamatan__kabupaten'
            )
        elif hasattr(self.request.user, 'role') and self.request.user.role == 'umkm':
            # UMKM can only see their own sales
            return ProdukTerjual.objects.filter(
                produk__umkm=self.request.user
            ).select_related(
                'produk',
                'produk__kategori',
                'lokasi_penjualan',
                'lokasi_penjualan__kecamatan',
                'lokasi_penjualan__kecamatan__kabupaten'
            )
        else:
            # Other users cannot see any sales
            return ProdukTerjual.objects.none()